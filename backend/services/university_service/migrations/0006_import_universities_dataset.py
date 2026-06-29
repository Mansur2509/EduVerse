"""One-shot data load of the committed university XLSX dataset.

This runs the existing, tested importer (`xlsx_import.import_rows`) exactly once
on deploy — Django records the migration as applied, so it does not re-run on
later deploys. It exists because the production environment has no shell access
to run the `import_universities_xlsx` management command manually.

Safety:
- Skipped entirely during test runs so it never loads 80 universities into every
  test database (which would break count-based tests).
- Wrapped in a savepoint + broad except: any parse/IO error rolls back the
  partial load and is logged, but never fails the migration / bricks the deploy.
- Idempotent: the importer upserts by slug, so re-loading (e.g. on a fresh
  environment replay) enriches rather than duplicates.
"""

from __future__ import annotations

import sys
from pathlib import Path

from django.db import migrations, transaction

SHEET = "Database"
WORKBOOK = Path(__file__).resolve().parents[3] / "data" / "universities" / "Universities Data.xlsx"


def load_universities(apps, schema_editor):
    # Never populate the test database; tests exercise the importer directly.
    if "test" in sys.argv:
        return

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[0006] openpyxl not installed; skipping university dataset import.")
        return

    if not WORKBOOK.exists():
        print(f"[0006] dataset workbook not found at {WORKBOOK}; skipping import.")
        return

    try:
        from services.university_service.xlsx_import import EXPECTED_HEADERS, import_rows

        workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
        if SHEET not in workbook.sheetnames:
            print(f"[0006] sheet {SHEET!r} not found; skipping import.")
            return
        worksheet = workbook[SHEET]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            print("[0006] dataset worksheet is empty; skipping import.")
            return

        index_by_header = {h: i for i, h in enumerate(rows[0]) if isinstance(h, str)}
        if any(header not in index_by_header for header in EXPECTED_HEADERS):
            print("[0006] dataset headers do not match the expected schema; skipping import.")
            return

        row_dicts = []
        for raw in rows[1:]:
            if not raw or all(cell in (None, "") for cell in raw):
                continue
            row_dicts.append(
                {h: (raw[i] if i < len(raw) else None) for h, i in index_by_header.items()}
            )

        with transaction.atomic():
            report = import_rows(row_dicts)
        summary = report.as_dict()["summary"]
        print(
            "[0006] university dataset import: "
            f"created={summary['created']} updated={summary['updated']} "
            f"skipped={summary['skipped']} warnings={summary['warnings']} "
            f"verifications={summary['fields_verified']}"
        )
    except Exception as exc:  # noqa: BLE001 - a data load must never brick a deploy
        print(f"[0006] university dataset import skipped due to error: {exc!r}")


class Migration(migrations.Migration):
    dependencies = [
        ("university_service", "0005_university_ap_recommendations_and_more"),
    ]

    operations = [
        migrations.RunPython(load_universities, migrations.RunPython.noop),
    ]
