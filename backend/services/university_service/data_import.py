"""Strict importer for the 72-column university admissions dataset.

This importer is deliberately separate from `xlsx_import.py`, which powers
the admin upload UI for another schema. The pipeline here is conservative:
every source cell is classified before it can reach a public field, existing
good data is not overwritten silently, and committed source rows are
fingerprinted so the same file cannot be imported twice by accident.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

from django.db import transaction

from .models import (
    University,
    UniversityDataImportBatch,
    UniversityDataImportRowLog,
    UniversityGuidanceContext,
    UniversitySignalWeights,
)

REQUIRED_COLUMNS = ("Name", "Country", "City")

CELL_STATUS_ACCEPTED = "accepted"
CELL_STATUS_NORMALIZED = "normalized"
IMPORTABLE_CELL_STATUSES = {CELL_STATUS_ACCEPTED, CELL_STATUS_NORMALIZED}

PLACEHOLDER_VALUES = {
    "",
    "-",
    "n/a",
    "na",
    "none",
    "not applicable",
    "not available",
    "not found",
    "not publicly available",
    "not used",
    "placeholder",
    "tba",
    "tbd",
    "unknown",
}

PLACEHOLDER_PATTERNS = (
    r"\bverify\b",
    r"\bverify exact\b",
    r"\bcheck official\b",
    r"\bsee\b.{0,40}\b(web)?site\b",
    r"\bconsult (the )?(official )?(web)?site\b",
    r"\bofficial catalogue\b",
    r"\bnot publicly available\b",
    r"\bnot found\b",
    r"\bplaceholder\b",
)

COMMENTARY_PATTERNS = (
    r"\bprogram list varies\b",
    r"\bvaries by faculty\b",
    r"\bvaries by programme\b",
    r"\bvaries by program\b",
    r"\bdepends on\b",
    r"\bdeadline varies\b",
    r"\bestimate only\b",
    r"\brough estimate\b",
    r"\bselection band\b",
    r"\bcompetitive by programme capacity\b",
    r"\bcompetitive by program capacity\b",
    r"\bavailable for eligible students\b",
)

GENERIC_COUNTRY_PATTERNS = (
    r"\baverage for this country\b",
    r"\bcountry average\b",
    r"\bin other system\b",
    r"\bconvert(ed|s|ing)?\b",
)

UNCERTAIN_PATTERNS = (
    r"\bmay vary\b",
    r"\bcan vary\b",
    r"\bapprox(imate|imately)?\b",
    r"\bnot confirmed\b",
)

URL_TRACKING_PARAMS = {"fbclid", "gclid", "mc_cid", "mc_eid"}
MONTH_RE = re.compile(
    r"\b("
    r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
    r"jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|"
    r"nov(?:ember)?|dec(?:ember)?"
    r")\b",
    re.IGNORECASE,
)
ROUND_RE = re.compile(
    r"\b(early action|early decision|regular decision|rd|ea|ed|rea|round [1-4]|fall|spring)\b",
    re.IGNORECASE,
)
URL_RE = re.compile(r"https?://[^\s)>,]+", re.IGNORECASE)
NUMBER_RE = re.compile(r"-?\d[\d,]*(?:\.\d+)?")
RANK_RE = re.compile(r"^\s*#?\s*(\d{1,5})(?:st|nd|rd|th)?(?:\s*(?:-|to)\s*(\d{1,5}))?\s*$", re.IGNORECASE)

PUBLIC_TEXT_FIELD_MAP = {
    "Official Website": "official_website",
    "Admissions URL": "admissions_url",
    "Admissions Website": "admissions_website",
    "Financial Aid Website": "financial_aid_url",
    "Application Portal": "application_portal_url",
    "International Students Office": "international_office_url",
    "Virtual Info Session": "virtual_info_session_url",
    "Deadlines": "deadlines_text",
    "Admissions Cycle Target": "admissions_cycle_target",
    "Standardized Testing Policy": "standardized_testing_policy_text",
    "Scholarships": "scholarships_text",
    "Need-based Aid": "need_based_aid_notes",
    "Merit Scholarship": "merit_scholarship_notes",
    "Other Scholarships": "other_scholarships_notes",
    "Scholarship Links": "scholarship_links_text",
    "AP Recommendations by Major": "ap_recommendations",
    "Application Requirements": "application_requirements",
    "Essays": "essay_requirements",
    "Profile Evidence": "profile_evidence_notes",
    "Activities": "activities_notes",
    "Honors / Olympiads": "honors_olympiads_notes",
    "Research Experience": "research_experience_notes",
    "Portfolio": "portfolio_notes",
    "Essay Drafts": "essay_drafts_notes",
}

URL_FIELDS = {
    "official_website",
    "admissions_url",
    "admissions_website",
    "financial_aid_url",
    "application_portal_url",
    "international_office_url",
    "virtual_info_session_url",
}

DEADLINE_COLUMNS = {"Deadlines"}
MAJOR_COLUMNS = {"Majors"}
SCHOLARSHIP_COLUMNS = {
    "Scholarships",
    "Need-based Aid",
    "Merit Scholarship",
    "Other Scholarships",
    "Scholarship Links",
}

NUMERIC_FIELD_CONFIG = {
    "IELTS Minimum": ("ielts_minimum", Decimal("0"), Decimal("9"), 1),
    "IELTS Competitive": ("ielts_competitive", Decimal("0"), Decimal("9"), 1),
    "Average GPA": ("gpa_average", Decimal("0"), Decimal("4.50"), 2),
    "QS Overall Score": ("qs_overall_score", Decimal("0"), Decimal("100"), 1),
    "Acceptance Rate": ("acceptance_rate", Decimal("0"), Decimal("100"), 2),
}

INTEGER_FIELD_CONFIG = {
    "SAT 25th": ("sat_p25", 400, 1600),
    "SAT 50th": ("sat_p50", 400, 1600),
    "SAT 75th": ("sat_p75", 400, 1600),
}

GUIDANCE_TEXT_FIELD_MAP = {
    "Recommendation Letters": "recommendation_letters",
    "What They Look For": "what_they_look_for",
    "Preferred Student Profile": "preferred_student_profile",
    "Who They Seek": "who_they_seek",
    "Student Traits Mentioned by University": "student_traits_mentioned",
    "Alumni Profile Evidence": "alumni_profile_evidence",
    "Published Admitted Student Essays": "published_admitted_student_essays",
    "Official Admissions Messaging": "official_admissions_messaging",
    "Student Life Page Signals": "student_life_page_signals",
    "Graduate/Alumni Outcomes": "graduate_alumni_outcomes",
    "Sample Admitted Essays": "sample_admitted_essays",
    "Essay Themes": "essay_themes",
    "Research/Leadership Themes": "research_leadership_themes",
    "Personality Traits Mentioned": "personality_traits_mentioned",
    "Academic Interests Mentioned": "academic_interests_mentioned",
    "Institutional Values": "institutional_values",
    "Source URLs": "source_urls",
    "Verification Status": "verification_status",
    "Data Source": "data_source",
    "Notes": "notes",
}

SIGNAL_SCORE_FIELD_MAP = {
    "Profile Evidence Score": "profile_evidence_score",
    "Activities Score": "activities_score",
    "Honors / Olympiads Score": "honors_olympiads_score",
    "Research Experience Score": "research_experience_score",
    "Portfolio Score": "portfolio_score",
    "Subject Passion Score": "subject_passion_score",
    "Curiosity Score": "curiosity_score",
    "Originality Score": "originality_score",
    "Leadership Score": "leadership_score",
    "Community Impact Score": "community_impact_score",
    "Research Fit Score": "research_fit_score",
    "Olympiads Score": "olympiads_score",
}

APPENDABLE_FIELDS = {
    "application_requirements",
    "ap_recommendations",
    "cost_notes",
    "data_quality_notes",
    "deadlines_text",
    "essay_drafts_notes",
    "essay_requirements",
    "financial_aid_notes",
    "honors_olympiads_notes",
    "merit_scholarship_notes",
    "need_based_aid_notes",
    "other_scholarships_notes",
    "portfolio_notes",
    "profile_evidence_notes",
    "research_experience_notes",
    "scholarship_links_text",
    "scholarships_text",
    "standardized_testing_policy_text",
    *GUIDANCE_TEXT_FIELD_MAP.values(),
}

ALIASES = {
    "eth zurich": "swiss federal institute technology zurich",
    "massachusetts institute of technology": "massachusetts institute technology",
    "mit": "massachusetts institute technology",
    "swiss federal institute of technology zurich": "swiss federal institute technology zurich",
    "ucl": "university college london",
    "university college london": "university college london",
}

AI_CLASSIFICATION_CACHE: dict[str, str] = {}


class ImportConfigurationError(Exception):
    """Raised for problems that stop the import before row processing."""


@dataclass(frozen=True)
class CleanedCell:
    raw_value: str
    cleaned_value: object | None
    status: str
    reason: str
    confidence: str

    @property
    def importable(self) -> bool:
        return self.status in IMPORTABLE_CELL_STATUSES


@dataclass
class AuditEntry:
    row_number: int
    university_name: str
    country: str
    matched_university_id: int | None
    action: str
    field_name: str
    raw_value: str
    cleaned_value: str
    status: str
    reason: str
    confidence: str


@dataclass
class ManualReviewEntry:
    row_number: int
    raw_name: str
    raw_country: str
    possible_matches: str
    conflict_fields: str
    raw_value: str
    existing_value: str
    new_cleaned_value: str
    reason: str


@dataclass
class RowResult:
    row_number: int
    name: str
    status: str
    row_hash: str = ""
    matched_university_id: int | None = None
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


@dataclass
class ImportSummary:
    rows_read: int = 0
    created: int = 0
    updated: int = 0
    skipped_existing: int = 0
    skipped_duplicate_rows: int = 0
    skipped_empty_rows: int = 0
    skipped_errors: int = 0
    duplicate_keys_in_file: int = 0
    missing_required: int = 0
    already_imported_rows: int = 0
    invalid_cells: int = 0
    placeholder_cells: int = 0
    commentary_cells: int = 0
    generic_country_average_cells: int = 0
    conflicts: int = 0
    ambiguous_matches: int = 0
    public_fields_imported: int = 0
    guidance_contexts_imported: int = 0
    signal_vectors_imported: int = 0
    accepted_cells_by_field: Counter = field(default_factory=Counter)
    skipped_cells_by_field: Counter = field(default_factory=Counter)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    rows: list[RowResult] = field(default_factory=list)
    audit_entries: list[AuditEntry] = field(default_factory=list)
    manual_review_entries: list[ManualReviewEntry] = field(default_factory=list)

    @property
    def updated_universities(self) -> int:
        return self.updated

    def add_cell_audit(self, entry: AuditEntry) -> None:
        self.audit_entries.append(entry)
        if entry.status in IMPORTABLE_CELL_STATUSES:
            self.accepted_cells_by_field[entry.field_name] += 1
            return
        self.skipped_cells_by_field[entry.field_name] += 1
        if entry.status == "skipped_placeholder":
            self.placeholder_cells += 1
        elif entry.status == "skipped_commentary":
            self.commentary_cells += 1
        elif entry.status == "skipped_generic_country_average":
            self.generic_country_average_cells += 1
        elif entry.status in {"skipped_wrong_field_type", "skipped_uncertain"}:
            self.invalid_cells += 1

    def add_manual_review(self, entry: ManualReviewEntry) -> None:
        self.manual_review_entries.append(entry)
        if entry.reason == "ambiguous_university_match":
            self.ambiguous_matches += 1
        else:
            self.conflicts += 1

    def to_summary_json(self) -> dict:
        return {
            "rows_read": self.rows_read,
            "created": self.created,
            "updated": self.updated,
            "skipped_duplicate_rows": self.skipped_duplicate_rows,
            "already_imported_rows": self.already_imported_rows,
            "invalid_cells": self.invalid_cells,
            "placeholder_cells": self.placeholder_cells,
            "commentary_cells": self.commentary_cells,
            "generic_country_average_cells": self.generic_country_average_cells,
            "conflicts": self.conflicts,
            "ambiguous_matches": self.ambiguous_matches,
            "accepted_cells_by_field": dict(self.accepted_cells_by_field),
            "skipped_cells_by_field": dict(self.skipped_cells_by_field),
        }


@dataclass(frozen=True)
class UniversityMatch:
    university: University | None
    reason: str
    possible_matches: list[University] = field(default_factory=list)

    @property
    def ambiguous(self) -> bool:
        return self.reason == "ambiguous"


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_placeholder(value: str) -> bool:
    return normalize_text(value) in PLACEHOLDER_VALUES


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", clean(value)).strip().lower()


def _matches_any(text: str, patterns: tuple[str, ...]) -> bool:
    lowered = normalize_text(text)
    return any(re.search(pattern, lowered, flags=re.IGNORECASE) for pattern in patterns)


def _base_skip_cell(raw_value, field_name: str) -> CleanedCell | None:
    raw = clean(raw_value)
    if not raw:
        return CleanedCell(raw, None, "skipped_empty", "empty cell", "high")
    if is_placeholder(raw):
        return CleanedCell(raw, None, "skipped_placeholder", "placeholder value", "high")
    if _matches_any(raw, GENERIC_COUNTRY_PATTERNS):
        return CleanedCell(
            raw,
            None,
            "skipped_generic_country_average",
            "generic country/conversion commentary, not university-specific data",
            "high",
        )
    if _matches_any(raw, PLACEHOLDER_PATTERNS):
        if field_name in {"Official Website", "Admissions URL", "Admissions Website"}:
            if URL_RE.search(raw):
                return None
        return CleanedCell(raw, None, "skipped_placeholder", "verification instruction", "high")
    if _matches_any(raw, COMMENTARY_PATTERNS):
        return CleanedCell(raw, None, "skipped_commentary", "commentary/generic note", "high")
    if _matches_any(raw, UNCERTAIN_PATTERNS):
        return CleanedCell(raw, None, "skipped_uncertain", "uncertain wording", "medium")
    return None


def classify_import_cell_with_ai(
    raw_value,
    field_name: str,
    university_name: str = "",
    country: str = "",
) -> str | None:
    """Optional AI-assisted stricter classifier.

    The current implementation intentionally does not call a provider. The
    environment flag and cache are present so a future backend-only provider
    adapter can plug in without changing importer semantics. Until then this
    returns None and deterministic validation remains authoritative.
    """

    if os.getenv("UNIVERSITY_IMPORT_AI_CLEANING_ENABLED", "false").lower() != "true":
        return None
    cache_key = hashlib.sha256(
        json.dumps(
            {
                "raw_value": clean(raw_value),
                "field_name": field_name,
                "university_name": university_name,
                "country": country,
            },
            sort_keys=True,
        ).encode("utf-8")
    ).hexdigest()
    return AI_CLASSIFICATION_CACHE.get(cache_key)


def _apply_ai_strictness(cell: CleanedCell, field_name: str, context: dict) -> CleanedCell:
    if not cell.importable:
        return cell
    classification = classify_import_cell_with_ai(
        cell.raw_value,
        field_name,
        context.get("name", ""),
        context.get("country", ""),
    )
    if classification in {None, "actual_university_data"}:
        return cell
    status = {
        "generic_country_average": "skipped_generic_country_average",
        "placeholder": "skipped_placeholder",
        "verification_instruction": "skipped_placeholder",
        "commentary": "skipped_commentary",
        "wrong_field_type": "skipped_wrong_field_type",
        "uncertain": "skipped_uncertain",
    }.get(classification, "skipped_uncertain")
    return CleanedCell(
        cell.raw_value,
        None,
        status,
        f"AI classifier marked cell as {classification}",
        "medium",
    )


def clean_raw_cell(value, field_name: str, university_context: dict | None = None) -> CleanedCell:
    context = university_context or {}
    raw = clean(value)
    base_skip = _base_skip_cell(raw, field_name)
    if base_skip:
        return base_skip

    if field_name in PUBLIC_TEXT_FIELD_MAP and PUBLIC_TEXT_FIELD_MAP[field_name] in URL_FIELDS:
        cell = _clean_url(raw)
    elif field_name in MAJOR_COLUMNS:
        cell = _clean_majors(raw)
    elif field_name in DEADLINE_COLUMNS:
        cell = _clean_deadline(raw)
    elif field_name in SCHOLARSHIP_COLUMNS:
        cell = _clean_scholarship_text(raw)
    elif field_name in NUMERIC_FIELD_CONFIG:
        attr, minimum, maximum, places = NUMERIC_FIELD_CONFIG[field_name]
        cell = _clean_decimal(raw, attr, minimum=minimum, maximum=maximum, places=places)
    elif field_name in INTEGER_FIELD_CONFIG:
        attr, minimum, maximum = INTEGER_FIELD_CONFIG[field_name]
        cell = _clean_int(raw, attr, minimum=minimum, maximum=maximum)
    elif field_name == "QS World University Ranking":
        cell = _clean_qs_rank(raw)
    elif field_name == "Tuition":
        cell = _clean_tuition(raw)
    elif field_name == "Last Verified Date":
        cell = _clean_date(raw)
    elif field_name in SIGNAL_SCORE_FIELD_MAP:
        cell = _clean_int(raw, field_name, minimum=0, maximum=10)
    else:
        cell = _clean_specific_text(raw)
    return _apply_ai_strictness(cell, field_name, context)


def _clean_url(text: str) -> CleanedCell:
    match = URL_RE.search(text)
    if not match:
        return CleanedCell(text, None, "skipped_wrong_field_type", "not an http(s) URL", "high")
    url = match.group(0).rstrip(".,;")
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return CleanedCell(text, None, "skipped_wrong_field_type", "invalid URL", "high")
    kept_query = [
        (key, value)
        for key, value in parse_qsl(parsed.query, keep_blank_values=True)
        if not key.lower().startswith("utm_") and key.lower() not in URL_TRACKING_PARAMS
    ]
    cleaned = urlunparse(
        (
            parsed.scheme,
            parsed.netloc.lower(),
            parsed.path or "/",
            "",
            urlencode(kept_query),
            "",
        )
    )
    status = CELL_STATUS_NORMALIZED if cleaned != text else CELL_STATUS_ACCEPTED
    return CleanedCell(text, cleaned, status, "valid http(s) URL", "high")


def _clean_decimal(
    text: str,
    field_name: str,
    *,
    minimum: Decimal,
    maximum: Decimal,
    places: int,
) -> CleanedCell:
    numbers = NUMBER_RE.findall(text.replace(",", ""))
    if not numbers:
        return CleanedCell(text, None, "skipped_wrong_field_type", "no numeric value", "high")
    allowed_words = {
        "acceptance_rate": ("%", "percent"),
        "gpa_average": ("/4", "gpa"),
        "ielts_minimum": ("ielts", "overall", "minimum"),
        "ielts_competitive": ("ielts", "overall", "competitive"),
        "qs_overall_score": ("qs", "score"),
    }.get(field_name, ())
    stripped = re.sub(r"[\d\s,./%<>~≈+-]", "", text.lower())
    for word in allowed_words:
        stripped = stripped.replace(word, "")
    if stripped.strip():
        return CleanedCell(text, None, "skipped_wrong_field_type", "prose in numeric field", "high")
    try:
        value = Decimal(numbers[0]).quantize(Decimal(1).scaleb(-places))
    except InvalidOperation:
        return CleanedCell(text, None, "skipped_wrong_field_type", "invalid decimal", "high")
    if value < minimum or value > maximum:
        return CleanedCell(text, None, "skipped_wrong_field_type", "number out of valid range", "high")
    return CleanedCell(text, value, CELL_STATUS_NORMALIZED, "validated numeric value", "high")


def _clean_int(text: str, field_name: str, *, minimum: int, maximum: int) -> CleanedCell:
    numbers = NUMBER_RE.findall(text.replace(",", ""))
    if len(numbers) != 1:
        return CleanedCell(text, None, "skipped_wrong_field_type", "expected one integer", "high")
    stripped = re.sub(r"[\d\s,./%<>~≈+-]", "", text.lower())
    allowed = ("sat", "score", "profile", "evidence", "activities")
    for word in allowed:
        stripped = stripped.replace(word, "")
    if stripped.strip():
        return CleanedCell(text, None, "skipped_wrong_field_type", "prose in integer field", "high")
    value = int(round(float(numbers[0])))
    if value < minimum or value > maximum:
        return CleanedCell(text, None, "skipped_wrong_field_type", "integer out of valid range", "high")
    return CleanedCell(text, value, CELL_STATUS_ACCEPTED, "validated integer", "high")


def _clean_qs_rank(text: str) -> CleanedCell:
    year_match = re.search(r"\b(20\d{2})\b", text)
    text_without_year = re.sub(r"\b20\d{2}\b", "", text)
    rank_match = RANK_RE.search(text_without_year.strip("# ,;()") or text)
    if not rank_match:
        ordinal = re.search(r"\b(\d{1,5})(?:st|nd|rd|th)\b", text, flags=re.IGNORECASE)
        rank = int(ordinal.group(1)) if ordinal else None
    else:
        rank = int(rank_match.group(1))
    if rank is None or rank <= 0:
        return CleanedCell(text, None, "skipped_wrong_field_type", "invalid QS rank", "high")
    return CleanedCell(
        text,
        {"rank": rank, "year": int(year_match.group(1)) if year_match else None},
        CELL_STATUS_NORMALIZED,
        "validated QS ranking",
        "high",
    )


def _clean_majors(text: str) -> CleanedCell:
    parts = re.split(r"[;\n]|(?<!\d),(?!\d)", text)
    seen: dict[str, str] = {}
    for part in parts:
        item = part.strip().strip(".")
        if not item or _base_skip_cell(item, "Majors"):
            continue
        if len(item.split()) > 12:
            continue
        key = normalize_text(item)
        seen.setdefault(key, item[:240])
    if not seen:
        return CleanedCell(text, None, "skipped_commentary", "no concrete major names", "high")
    return CleanedCell(text, list(seen.values()), CELL_STATUS_NORMALIZED, "deduped major list", "high")


def _clean_deadline(text: str) -> CleanedCell:
    has_date = bool(
        MONTH_RE.search(text)
        or re.search(r"\b\d{1,2}[/-]\d{1,2}(?:[/-]\d{2,4})?\b", text)
        or ROUND_RE.search(text)
    )
    if not has_date:
        return CleanedCell(text, None, "skipped_uncertain", "no concrete deadline signal", "medium")
    return CleanedCell(text, _dedupe_text_chunks("", text), CELL_STATUS_ACCEPTED, "specific deadline text", "medium")


def _clean_scholarship_text(text: str) -> CleanedCell:
    has_specific_signal = bool(
        re.search(r"\b(scholarship|fellowship|grant|aid|tuition|merit|need-based)\b", text, re.I)
        and not re.fullmatch(r"available for eligible students\.?", normalize_text(text))
    )
    has_amount_or_url = bool(URL_RE.search(text) or re.search(r"[$€£¥]\s?\d|\b\d+%", text))
    if not (has_specific_signal or has_amount_or_url):
        return CleanedCell(text, None, "skipped_uncertain", "generic scholarship text", "medium")
    return CleanedCell(text, _dedupe_text_chunks("", text), CELL_STATUS_ACCEPTED, "specific aid text", "medium")


def _clean_tuition(text: str) -> CleanedCell:
    if not re.search(r"[$€£¥]\s?\d|\b(usd|eur|gbp|cad|aud|chf|jpy|sgd)\b|\d", text, re.I):
        return CleanedCell(text, None, "skipped_wrong_field_type", "no tuition amount/currency", "high")
    amount = None
    match = NUMBER_RE.search(text.replace(",", ""))
    if match:
        try:
            amount = Decimal(match.group()).quantize(Decimal("0.01"))
        except InvalidOperation:
            amount = None
    return CleanedCell(
        text,
        {"amount": amount, "notes": text[:4000]},
        CELL_STATUS_NORMALIZED,
        "validated tuition text",
        "medium",
    )


def _clean_date(text: str) -> CleanedCell:
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%B %d, %Y", "%b %d, %Y"):
        try:
            return CleanedCell(text, datetime.strptime(text, fmt).date(), CELL_STATUS_NORMALIZED, "parsed date", "high")
        except ValueError:
            continue
    if re.fullmatch(r"\d{4,6}(\.0+)?", text):
        try:
            serial = int(float(text))
            return CleanedCell(
                text,
                date(1899, 12, 30) + timedelta(days=serial),
                CELL_STATUS_NORMALIZED,
                "parsed Excel serial date",
                "medium",
            )
        except (ValueError, OverflowError):
            pass
    return CleanedCell(text, None, "skipped_wrong_field_type", "invalid date", "high")


def _clean_specific_text(text: str) -> CleanedCell:
    if len(text) < 3:
        return CleanedCell(text, None, "skipped_uncertain", "too short to be useful", "medium")
    return CleanedCell(text, _dedupe_text_chunks("", text), CELL_STATUS_ACCEPTED, "specific text", "medium")


def clean_public_text(value) -> str:
    cell = clean_raw_cell(value, "Text")
    return str(cell.cleaned_value or "") if cell.importable else ""


def looks_like_http_url(value: str) -> bool:
    return _clean_url(value).importable


def parse_optional_int(value) -> tuple[int | None, str | None]:
    base_skip = _base_skip_cell(value, "integer")
    if base_skip:
        if base_skip.status in {"skipped_empty", "skipped_placeholder"}:
            return None, None
        return None, base_skip.reason
    cell = _clean_int(clean(value), "integer", minimum=0, maximum=999999)
    if cell.status in {"skipped_empty", "skipped_placeholder"}:
        return None, None
    if not cell.importable:
        return None, cell.reason
    return int(cell.cleaned_value), None


def parse_optional_decimal(value, *, max_digits: int, decimal_places: int) -> tuple[Decimal | None, str | None]:
    base_skip = _base_skip_cell(value, "decimal")
    if base_skip:
        if base_skip.status in {"skipped_empty", "skipped_placeholder"}:
            return None, None
        return None, base_skip.reason
    maximum = Decimal(10) ** (max_digits - decimal_places) - Decimal(1).scaleb(-decimal_places)
    cell = _clean_decimal(
        clean(value),
        "decimal",
        minimum=Decimal("-999999"),
        maximum=maximum,
        places=decimal_places,
    )
    if cell.status == "skipped_empty" or cell.status == "skipped_placeholder":
        return None, None
    if not cell.importable:
        return None, cell.reason
    return cell.cleaned_value, None


def parse_score_0_10(value) -> tuple[int | None, str | None]:
    base_skip = _base_skip_cell(value, "score")
    if base_skip:
        if base_skip.status in {"skipped_empty", "skipped_placeholder"}:
            return None, None
        return None, base_skip.reason
    cell = _clean_int(clean(value), "score", minimum=0, maximum=10)
    if cell.status == "skipped_empty" or cell.status == "skipped_placeholder":
        return None, None
    if not cell.importable:
        return None, cell.reason
    return int(cell.cleaned_value), None


def parse_date_safe(value) -> tuple[date | None, str | None]:
    cell = clean_raw_cell(value, "Last Verified Date")
    if cell.status == "skipped_empty" or cell.status == "skipped_placeholder":
        return None, None
    if not cell.importable:
        return None, cell.reason
    return cell.cleaned_value, None


def parse_qs_ranking(value) -> tuple[int | None, int | None, str | None]:
    cell = clean_raw_cell(value, "QS World University Ranking")
    if cell.status == "skipped_empty" or cell.status == "skipped_placeholder":
        return None, None, None
    if not cell.importable:
        return None, None, cell.reason
    payload = cell.cleaned_value or {}
    return payload.get("rank"), payload.get("year"), None


def split_majors(value) -> list[str]:
    cell = clean_raw_cell(value, "Majors")
    return list(cell.cleaned_value or []) if cell.importable else []


def _canonical_name(value: str) -> str:
    raw = re.sub(r"\s*\([^)]*\)\s*$", "", clean(value))
    raw = raw.lower().replace("&", " and ")
    raw = re.sub(r"[^a-z0-9\s]", " ", raw)
    raw = re.sub(r"\bthe\b", " ", raw)
    raw = re.sub(r"\b(university|uni)\b$", " ", raw)
    raw = raw.replace("institute of technology", "institute technology")
    raw = re.sub(r"\s+", " ", raw).strip()
    return ALIASES.get(raw, raw)


def _canonical_country(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", clean(value).lower()).strip()


def normalized_university_key(name: str, country: str) -> tuple[str, str]:
    return _canonical_name(name), _canonical_country(country)


def strip_parenthetical(name: str) -> str:
    return re.sub(r"\s*\([^)]*\)\s*$", "", name).strip()


def normalize_key_part(value: str) -> str:
    return normalize_text(value)


def _domain_from_url(value: str) -> str:
    cell = _clean_url(value)
    if not cell.importable:
        return ""
    host = urlparse(str(cell.cleaned_value)).netloc.lower()
    return host[4:] if host.startswith("www.") else host


def _row_hash(row: dict) -> str:
    normalized = {
        clean(key): normalize_text(value)
        for key, value in sorted(row.items())
        if clean(key) and clean(value)
    }
    return hashlib.sha256(json.dumps(normalized, sort_keys=True).encode("utf-8")).hexdigest()


def _dedupe_text_chunks(existing: str, new_text: str) -> str:
    chunks = re.split(r"(?:\n+|;\s+|(?<=[.!?])\s+)", new_text)
    existing_chunks = [
        chunk.strip()
        for chunk in re.split(r"(?:\n+|;\s+|(?<=[.!?])\s+)", existing or "")
        if chunk.strip()
    ]
    seen = {normalize_text(chunk) for chunk in existing_chunks}
    additions = []
    for chunk in chunks:
        cleaned = chunk.strip()
        key = normalize_text(cleaned)
        if cleaned and key not in seen:
            additions.append(cleaned)
            seen.add(key)
    return "\n".join([*existing_chunks, *additions]).strip()


def read_rows(path: str) -> tuple[list[str], list[dict[str, str]]]:
    file_path = Path(path)
    if not file_path.is_file():
        raise ImportConfigurationError(f"File not found: {path}")

    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(file_path)
    if suffix in (".csv", ".tsv", ".txt"):
        return _read_delimited(file_path)
    raise ImportConfigurationError(f"Unsupported file extension: {suffix}")


def _read_xlsx(file_path: Path) -> tuple[list[str], list[dict[str, str]]]:
    try:
        import openpyxl
    except ImportError as error:  # pragma: no cover
        raise ImportConfigurationError("openpyxl is required to read .xlsx files.") from error

    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet = workbook["University Data"] if "University Data" in workbook.sheetnames else workbook.worksheets[0]
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], []
        headers = [clean(cell) for cell in header_row]
        rows = []
        for raw_row in rows_iter:
            if raw_row is None or all(cell is None or clean(cell) == "" for cell in raw_row):
                continue
            row = {headers[i]: raw_row[i] if i < len(raw_row) else None for i in range(len(headers))}
            rows.append(row)
        return headers, rows
    finally:
        workbook.close()


def _read_delimited(file_path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with open(file_path, encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = "\t" if file_path.suffix.lower() == ".tsv" else ","
        reader = csv.DictReader(handle, dialect=dialect)
        headers = [clean(h) for h in (reader.fieldnames or [])]
        rows = [row for row in reader if any(clean(value) for value in row.values())]
        return headers, rows


def _build_fields(row: dict, row_number: int, context: dict, summary: ImportSummary) -> tuple[dict, dict, dict]:
    public_fields: dict = {}
    guidance_fields: dict = {}
    signal_fields: dict = {}
    raw_context: dict = {}

    def add_audit(column: str, cell: CleanedCell, action: str = "validate") -> None:
        summary.add_cell_audit(
            AuditEntry(
                row_number=row_number,
                university_name=context["name"],
                country=context["country"],
                matched_university_id=None,
                action=action,
                field_name=column,
                raw_value=cell.raw_value,
                cleaned_value=_stringify(cell.cleaned_value),
                status=cell.status,
                reason=cell.reason,
                confidence=cell.confidence,
            )
        )

    for column, attr in PUBLIC_TEXT_FIELD_MAP.items():
        cell = clean_raw_cell(row.get(column), column, context)
        if cell.status != "skipped_empty":
            add_audit(column, cell)
        if not cell.importable:
            continue
        public_fields[attr] = cell.cleaned_value

    majors_cell = clean_raw_cell(row.get("Majors"), "Majors", context)
    if majors_cell.status != "skipped_empty":
        add_audit("Majors", majors_cell)
    if majors_cell.importable:
        public_fields["majors_list"] = majors_cell.cleaned_value

    for column, (attr, _minimum, _maximum, _places) in NUMERIC_FIELD_CONFIG.items():
        cell = clean_raw_cell(row.get(column), column, context)
        if cell.status != "skipped_empty":
            add_audit(column, cell)
        if cell.importable:
            public_fields[attr] = cell.cleaned_value

    for column, (attr, _minimum, _maximum) in INTEGER_FIELD_CONFIG.items():
        cell = clean_raw_cell(row.get(column), column, context)
        if cell.status != "skipped_empty":
            add_audit(column, cell)
        if cell.importable:
            public_fields[attr] = cell.cleaned_value

    qs_cell = clean_raw_cell(row.get("QS World University Ranking"), "QS World University Ranking", context)
    if qs_cell.status != "skipped_empty":
        add_audit("QS World University Ranking", qs_cell)
    if qs_cell.importable:
        payload = qs_cell.cleaned_value
        public_fields["qs_ranking"] = payload["rank"]
        if payload.get("year"):
            public_fields["qs_ranking_year"] = payload["year"]

    tuition_cell = clean_raw_cell(row.get("Tuition"), "Tuition", context)
    if tuition_cell.status != "skipped_empty":
        add_audit("Tuition", tuition_cell)
    if tuition_cell.importable:
        payload = tuition_cell.cleaned_value
        if payload.get("amount") is not None:
            public_fields["tuition_amount"] = payload["amount"]
        public_fields["cost_notes"] = payload["notes"]

    for column, attr in GUIDANCE_TEXT_FIELD_MAP.items():
        raw_value = clean(row.get(column))
        if raw_value:
            raw_context[column] = raw_value
        cell = clean_raw_cell(row.get(column), column, context)
        if cell.status != "skipped_empty":
            add_audit(column, cell)
        if cell.importable:
            guidance_fields[attr] = cell.cleaned_value

    date_cell = clean_raw_cell(row.get("Last Verified Date"), "Last Verified Date", context)
    if date_cell.status != "skipped_empty":
        add_audit("Last Verified Date", date_cell)
    if date_cell.importable:
        guidance_fields["last_verified_date"] = date_cell.cleaned_value

    guidance_fields["raw_context_json"] = raw_context

    for column, attr in SIGNAL_SCORE_FIELD_MAP.items():
        cell = clean_raw_cell(row.get(column), column, context)
        if cell.status != "skipped_empty":
            add_audit(column, cell)
        if cell.importable:
            signal_fields[attr] = cell.cleaned_value

    source_cell = clean_raw_cell(row.get("Profile Scoring Source"), "Profile Scoring Source", context)
    if source_cell.status != "skipped_empty":
        add_audit("Profile Scoring Source", source_cell)
    if source_cell.importable:
        signal_fields["profile_scoring_source"] = source_cell.cleaned_value

    return public_fields, guidance_fields, signal_fields


def _stringify(value) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, list | dict):
        return json.dumps(_json_safe(value), ensure_ascii=False, sort_keys=True)
    return str(value)


def _json_safe(value):
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if isinstance(value, dict):
        return {key: _json_safe(item) for key, item in value.items()}
    return value


def _build_existing_indexes() -> tuple[dict[tuple[str, str], University], dict[str, list[University]], list[University]]:
    exact: dict[tuple[str, str], University] = {}
    by_domain: dict[str, list[University]] = {}
    universities = list(University.objects.select_related("guidance_context", "signal_weights").all())
    for university in universities:
        exact[normalized_university_key(university.name, university.country)] = university
        domain = _domain_from_url(university.official_website)
        if domain:
            by_domain.setdefault(domain, []).append(university)
    return exact, by_domain, universities


def _match_university(
    name: str,
    country: str,
    website: str,
    exact: dict[tuple[str, str], University],
    by_domain: dict[str, list[University]],
    universities: list[University],
) -> UniversityMatch:
    key = normalized_university_key(name, country)
    if key in exact:
        return UniversityMatch(exact[key], "exact")
    domain = _domain_from_url(website)
    if domain and domain in by_domain:
        matches = by_domain[domain]
        if len(matches) == 1:
            return UniversityMatch(matches[0], "domain")
        return UniversityMatch(None, "ambiguous", matches)
    candidates = [
        university
        for university in universities
        if _canonical_country(university.country) == key[1]
    ]
    scored = [
        (SequenceMatcher(None, key[0], normalized_university_key(candidate.name, candidate.country)[0]).ratio(), candidate)
        for candidate in candidates
    ]
    strong = [(score, candidate) for score, candidate in scored if score >= 0.94]
    if len(strong) == 1:
        return UniversityMatch(strong[0][1], "fuzzy")
    if len(strong) > 1:
        return UniversityMatch(None, "ambiguous", [candidate for _score, candidate in strong])
    return UniversityMatch(None, "new")


def _is_missing_existing(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip() or is_placeholder(value)
    if isinstance(value, list | dict):
        return not value
    return False


def _values_equal(existing, new_value) -> bool:
    if isinstance(existing, Decimal) or isinstance(new_value, Decimal):
        try:
            return Decimal(str(existing)) == Decimal(str(new_value))
        except InvalidOperation:
            return False
    if isinstance(existing, list) or isinstance(new_value, list):
        return [normalize_text(item) for item in (existing or [])] == [
            normalize_text(item) for item in (new_value or [])
        ]
    return normalize_text(existing) == normalize_text(new_value)


def _apply_fields(
    target,
    fields: dict,
    *,
    row_number: int,
    raw_name: str,
    raw_country: str,
    summary: ImportSummary,
    update_existing: bool,
    model_label: str,
) -> int:
    changed = 0
    for attr, new_value in fields.items():
        if attr == "raw_context_json":
            existing = getattr(target, attr, {})
            merged = {**(existing or {}), **(new_value or {})}
            if merged != existing:
                setattr(target, attr, merged)
                changed += 1
            continue
        existing = getattr(target, attr, None)
        if _is_missing_existing(existing):
            setattr(target, attr, new_value)
            changed += 1
            continue
        if _values_equal(existing, new_value):
            continue
        if update_existing and attr in APPENDABLE_FIELDS:
            merged = _dedupe_text_chunks(str(existing), str(new_value))
            if normalize_text(merged) != normalize_text(existing):
                setattr(target, attr, merged)
                changed += 1
            continue
        summary.add_manual_review(
            ManualReviewEntry(
                row_number=row_number,
                raw_name=raw_name,
                raw_country=raw_country,
                possible_matches=str(getattr(target, "id", "")),
                conflict_fields=f"{model_label}.{attr}",
                raw_value="",
                existing_value=_stringify(existing),
                new_cleaned_value=_stringify(new_value),
                reason="conflict",
            )
        )
    return changed


def _unique_slug(name: str) -> str:
    from django.utils.text import slugify

    base = slugify(strip_parenthetical(name) or name)[:250] or "university"
    slug = base
    suffix = 2
    while University.objects.filter(slug=slug).exists():
        slug = f"{base}-{suffix}"[:260]
        suffix += 1
    return slug


def _process_row(
    row_number: int,
    row: dict,
    *,
    batch: UniversityDataImportBatch | None,
    existing_exact: dict[tuple[str, str], University],
    existing_domains: dict[str, list[University]],
    existing_universities: list[University],
    seen_keys: set[tuple[str, str]],
    update_existing: bool,
    force_reprocess: bool,
    source_file_name: str,
    summary: ImportSummary,
) -> None:
    name = clean(row.get("Name"))
    country = clean(row.get("Country"))
    city = clean(row.get("City"))
    row_hash = _row_hash(row)

    if not any(clean(value) for value in row.values()):
        summary.skipped_empty_rows += 1
        summary.rows.append(RowResult(row_number, "(blank)", "skipped_empty", row_hash=row_hash))
        return
    if not name or not country or not city:
        summary.missing_required += 1
        summary.skipped_errors += 1
        message = f"row {row_number}: missing required Name/Country/City"
        summary.errors.append(message)
        summary.rows.append(RowResult(row_number, name or "(blank)", "skipped_error", row_hash=row_hash, errors=[message]))
        return
    if (
        not force_reprocess
        and UniversityDataImportRowLog.objects.filter(row_hash=row_hash, batch__committed=True).exists()
    ):
        summary.already_imported_rows += 1
        summary.skipped_duplicate_rows += 1
        summary.rows.append(RowResult(row_number, name, "already_imported", row_hash=row_hash))
        return

    key = normalized_university_key(name, country)
    if key in seen_keys:
        summary.duplicate_keys_in_file += 1
        summary.skipped_duplicate_rows += 1
        summary.rows.append(RowResult(row_number, name, "skip_duplicate", row_hash=row_hash))
        return
    seen_keys.add(key)

    context = {"name": name, "country": country}
    public_fields, guidance_fields, signal_fields = _build_fields(row, row_number, context, summary)
    match = _match_university(
        name,
        country,
        str(public_fields.get("official_website") or row.get("Official Website") or ""),
        existing_exact,
        existing_domains,
        existing_universities,
    )
    if match.ambiguous:
        summary.add_manual_review(
            ManualReviewEntry(
                row_number=row_number,
                raw_name=name,
                raw_country=country,
                possible_matches="; ".join(f"{item.id}:{item.name}" for item in match.possible_matches),
                conflict_fields="",
                raw_value="",
                existing_value="",
                new_cleaned_value="",
                reason="ambiguous_university_match",
            )
        )
        summary.rows.append(RowResult(row_number, name, "manual_review", row_hash=row_hash))
        return

    existing = match.university
    if existing is None:
        university = University(
            name=name,
            country=country,
            city=city,
            slug=_unique_slug(name),
            is_published=True,
        )
        for attr, value in public_fields.items():
            setattr(university, attr, value)
        university.save()
        UniversityGuidanceContext.objects.create(university=university, **guidance_fields)
        UniversitySignalWeights.objects.create(university=university, **signal_fields)
        summary.created += 1
        summary.public_fields_imported += len(public_fields)
        summary.guidance_contexts_imported += 1
        summary.signal_vectors_imported += 1
        existing_exact[key] = university
        existing_universities.append(university)
        action = "create"
    else:
        public_changed = _apply_fields(
            existing,
            {**({"city": city} if city else {}), **public_fields},
            row_number=row_number,
            raw_name=name,
            raw_country=country,
            summary=summary,
            update_existing=update_existing,
            model_label="University",
        )
        if public_changed:
            existing.save()
        guidance, _ = UniversityGuidanceContext.objects.get_or_create(university=existing)
        guidance_changed = _apply_fields(
            guidance,
            guidance_fields,
            row_number=row_number,
            raw_name=name,
            raw_country=country,
            summary=summary,
            update_existing=update_existing,
            model_label="UniversityGuidanceContext",
        )
        if guidance_changed:
            guidance.save()
        signals, _ = UniversitySignalWeights.objects.get_or_create(university=existing)
        signal_changed = _apply_fields(
            signals,
            signal_fields,
            row_number=row_number,
            raw_name=name,
            raw_country=country,
            summary=summary,
            update_existing=update_existing,
            model_label="UniversitySignalWeights",
        )
        if signal_changed:
            signals.save()
        changed = public_changed + guidance_changed + signal_changed
        if changed:
            summary.updated += 1
            summary.public_fields_imported += public_changed
            summary.guidance_contexts_imported += 1 if guidance_changed else 0
            summary.signal_vectors_imported += 1 if signal_changed else 0
            action = "update_missing"
        else:
            summary.skipped_existing += 1
            action = "skip_duplicate"
        university = existing

    for entry in summary.audit_entries:
        if entry.row_number == row_number:
            entry.matched_university_id = university.id
            entry.action = action
    if batch is not None:
        UniversityDataImportRowLog.objects.create(
            batch=batch,
            source_file_name=source_file_name,
            row_number=row_number,
            row_hash=row_hash,
            matched_university=university,
            action=action,
        )
    summary.rows.append(
        RowResult(row_number, name, action, row_hash=row_hash, matched_university_id=university.id)
    )


def import_universities_data(
    path: str,
    *,
    commit: bool = False,
    update_existing: bool = False,
    missing_only: bool = True,
    limit: int | None = None,
    audit_out: str | None = None,
    manual_review_out: str | None = None,
    force_reprocess: bool = False,
) -> ImportSummary:
    """Clean, dedupe, and safely upsert the university data file."""

    headers, rows = read_rows(path)
    missing_columns = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing_columns:
        raise ImportConfigurationError(f"Missing required column(s): {', '.join(missing_columns)}")
    if limit is not None:
        rows = rows[:limit]

    source_file_name = Path(path).name
    summary = ImportSummary(rows_read=len(rows))
    existing_exact, existing_domains, existing_universities = _build_existing_indexes()
    seen_keys: set[tuple[str, str]] = set()
    batch: UniversityDataImportBatch | None = None

    def run() -> None:
        nonlocal batch
        if commit:
            batch = UniversityDataImportBatch.objects.create(
                source_file_name=source_file_name,
                committed=True,
                row_count=len(rows),
            )
        for index, row in enumerate(rows, start=2):
            try:
                _process_row(
                    index,
                    row,
                    batch=batch,
                    existing_exact=existing_exact,
                    existing_domains=existing_domains,
                    existing_universities=existing_universities,
                    seen_keys=seen_keys,
                    update_existing=update_existing and not missing_only,
                    force_reprocess=force_reprocess,
                    source_file_name=source_file_name,
                    summary=summary,
                )
            except Exception as error:  # noqa: BLE001 - row-level isolation
                summary.skipped_errors += 1
                message = f"row {index}: unexpected error ({type(error).__name__}): {error}"
                summary.errors.append(message)
                summary.rows.append(
                    RowResult(index, clean(row.get("Name")) or "(unknown)", "skipped_error", errors=[message])
                )
        if batch is not None:
            batch.summary_json = summary.to_summary_json()
            batch.save(update_fields=["summary_json"])

    with transaction.atomic():
        run()
        if not commit:
            transaction.set_rollback(True)

    if audit_out:
        write_audit_csv(summary, audit_out)
    if manual_review_out:
        write_manual_review_csv(summary, manual_review_out)
    return summary


def write_audit_csv(summary: ImportSummary, path: str) -> None:
    fieldnames = [
        "row_number",
        "university_name",
        "country",
        "matched_university_id",
        "action",
        "field_name",
        "raw_value",
        "cleaned_value",
        "status",
        "reason",
        "confidence",
    ]
    with open(path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for entry in summary.audit_entries:
            writer.writerow(asdict(entry))


def write_manual_review_csv(summary: ImportSummary, path: str) -> None:
    fieldnames = [
        "row_number",
        "raw_name",
        "raw_country",
        "possible_matches",
        "conflict_fields",
        "raw_value",
        "existing_value",
        "new_cleaned_value",
        "reason",
    ]
    with open(path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for entry in summary.manual_review_entries:
            writer.writerow(asdict(entry))
